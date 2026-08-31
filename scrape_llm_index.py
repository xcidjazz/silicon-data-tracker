"""Daily scraper for Silicon Data indices.

Datasets captured (each becomes one sheet in one workbook, one CSV archive):
  1. LLM token expenditure  - LLM Token / Open LLM / Proprietary LLM  (4dp)
  2. GPU rental prices      - H100/A100/H200/MI300X neo-cloud, H100/A100 hyperscaler (2dp)

WORKBOOK OWNERSHIP CONTRACT - read before editing the workbook by hand:

  Machine-owned (appended to or regenerated every run, do not hand-edit):
    * sheet 'Daily Index'  columns A:G   - LLM token indices
    * sheet 'GPU Rental'   columns A:O   - GPU rental indices
    * sheets 'Chart' and 'Source'        - rebuilt each run

  Yours (the script never touches these):
    * 'Daily Index' columns H onward, 'GPU Rental' columns P onward
    * any other sheet you create

Rows are APPENDED. Existing rows are never rewritten, because each endpoint only
exposes a rolling 7-day window - older values are gone from the web forever.
The CSVs under data/ are the permanent archive and the ultimate source of truth:
the workbook is reconstructible from them, they are not reconstructible from the
web. Both are snapshotted to backups/ before every write.

Run:  python scrape_llm_index.py [--no-xlsx] [--quiet] [--rebuild] [--only llm|gpu]
"""

import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
from datetime import date, datetime, timezone
from pathlib import Path

import requests

BASE = Path(__file__).resolve().parent
XLSX_PATH = BASE / "LLM_Token_Expenditure_Index.xlsx"
LOG_PATH = BASE / "logs" / "scrape.log"
BACKUP_DIR = BASE / "backups"
KEEP_BACKUPS = 30
TELEGRAM = Path.home() / ".claude" / "scripts" / "telegram_notify.py"

CHART_SHEET = "Chart"
SOURCE_SHEET = "Source"

TOKEN_PORTAL = "https://portal.silicondata.com/token-indexes-chart"
GPU_PORTAL = "https://portal.silicondata.com/gpu-index-chart"
MARKETING = "https://www.silicondata.com/products/silicon-index/llm-token-expenditure-index"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36")

# ---------------------------------------------------------------- LLM dataset
# portal token key -> (csv column, expected index_name, marketing readings key)
LLM_SERIES = {
    "expenditure": ("llm_token", "LLM Token Expenditure Index", "llm-token"),
    "open_expenditure": ("open_llm", "Open LLM Token Expenditure Index", "open-llm"),
    "closed_expenditure": ("proprietary_llm", "Proprietary LLM Token Expenditure Index", "proprietary-llm"),
}

# ---------------------------------------------------------------- GPU dataset
# Only these 7 combinations exist. Confirmed against the site's own chart config:
#   hyperscaler: [H100, A100]
#   neo-cloud:   [H100, A100, H200, MI300X, B200]
# Requesting mainTab=hyperscaler for B200/MI300X/H200 does NOT error - the server
# silently returns the NEO-CLOUD series while echoing initialMainTab="neo-cloud".
# Scraping those blindly would write neo-cloud numbers into hyperscaler columns,
# so fetch_gpu asserts the echoed tab matches what was requested.
# H200 is hidden from the public site (filtered unless standalone=true) but the
# series is real and populated; captured because a day not captured is lost forever.
GPU_SERIES = [
    ("h100_neo", "h100", "neo-cloud", "SDH100RT"),
    ("a100_neo", "a100", "neo-cloud", "SDA100RT"),
    ("h200_neo", "h200", "neo-cloud", ""),
    ("mi300x_neo", "mi300x", "neo-cloud", "SDMI300XRT"),
    ("b200_neo", "b200", "neo-cloud", "SDB200RT"),
    ("h100_hyper", "h100", "hyperscaler", ""),
    ("a100_hyper", "a100", "hyperscaler", ""),
]

# ------------------------------------------------------------ forward curves
# One snapshot per day, no history and no backfill (confirmed: date params are
# ignored, the response is always a single curve). A day not captured is lost.
# The payload carries 145 tenors per GPU: 37 integer months (0-36) plus 108
# quarter-month points that are interpolation for smooth chart rendering.
# The sheet and CSV keep the 37 integer months - enough to draw the curve and
# sustainable in Excel - while the FULL 145-tenor payload is archived verbatim
# as daily JSON so nothing is ever actually discarded.
FC_GPUS = ["H100", "A100", "B200"]
FC_RATES = ["term_rate", "forward_rate"]
FC_TENORS = [str(t) for t in range(0, 37)]
FC_SHEET = "Forward Curve"
FC_CSV = BASE / "data" / "forward_curve.csv"
FC_RAW_DIR = BASE / "data" / "forward_curve_raw"
FC_PORTAL = "https://portal.silicondata.com/forward-curve-chart"
FC_COLUMNS = ["date", "gpu", "rate_type"] + [f"t{t}" for t in FC_TENORS]

# ---------------------------------------------------------- Ramp AI Index
# Monthly, published with a lag (July data as of late August). Unlike the Silicon
# Data endpoints this serves the FULL history every time, so it self-heals and a
# missed day costs nothing - we just check daily and append when a new month lands.
# Field mapping verified against Citadel Securities' published July figures:
# top 1% +48.8% (they printed 49%), top 10% +24.9% (25%), median +9.3% (9%).
# Note Citadel's "90th percentile" is this feed's top_10_percent_median_pepm
# (650/11.95 = 54x, their quoted figure), NOT p90_pepm.
RAMP_URL = "https://ramp.com/data/ai-index"
RAMP_RAW_DIR = BASE / "data" / "ramp_raw"

# (breakdown label, source array, dimension field, dimension_type filter)
RAMP_ADOPTION = [
    ("Headline", "adoptionOverall", None, None),
    # The site's headline chart plots this dashed alongside Headline: Ramp's estimate
    # of adoption across all US businesses, not just firms on Ramp. Much lower level
    # (21.6% vs 55.7%) because the Ramp base skews to earlier adopters.
    ("US estimate", "adoptionUsEstimate", None, None),
    ("Sector", "adoptionIndustry", "naics_sector", None),
    ("Business size", "adoptionSize", "business_size", None),
    ("Models", "adoptionVendor", "vendor", None),
]
RAMP_SPEND = [
    ("Overall", "spendPerEmployee", None, None),
    ("Business size", "spendPerEmployeeCurated", "dimension_label", "fte_segment"),
    ("Sector", "spendPerEmployeeCurated", "dimension_label", "naics_sector"),
    ("Financing status", "spendPerEmployeeCurated", "dimension_label", "company_financing_status"),
]
RAMP_ADOPTION_METRICS = ["adoption_rate_pct", "mom_change_pp", "yoy_change_pp"]
RAMP_SPEND_METRICS = ["median_pepm", "p90_pepm", "p99_pepm", "top_10_percent_median_pepm",
                      "top_1_percent_median_pepm", "p99_winsorized_weighted_pepm"]
RAMP_SHEETS = {
    "ramp_adoption": {
        "sheet": "Ramp Adoption",
        "csv": BASE / "data" / "ramp_adoption.csv",
        "specs": RAMP_ADOPTION,
        "metrics": RAMP_ADOPTION_METRICS,
        "labels": ["Month", "Breakdown", "Dimension", "Adoption %", "MoM (pp)", "YoY (pp)"],
        "widths": [12, 15, 44, 12, 11, 11],
        "fmt": "0.00",
    },
    "ramp_spend": {
        "sheet": "Ramp Spend per Employee",
        "csv": BASE / "data" / "ramp_spend.csv",
        "specs": RAMP_SPEND,
        "metrics": RAMP_SPEND_METRICS,
        "labels": ["Month", "Breakdown", "Dimension", "Median $", "P90 $", "P99 $",
                   "Top 10% median $", "Top 1% median $", "Winsorized wtd $"],
        "widths": [12, 17, 44, 12, 12, 12, 17, 17, 17],
        "fmt": "#,##0.00",
    },
}

DATASETS = {
    "llm": {
        "sheet": "Daily Index",
        "csv": BASE / "data" / "llm_token_index.csv",
        "columns": ["date", "llm_token", "open_llm", "proprietary_llm"],
        "labels": ["Date", "LLM Token", "Open LLM", "Proprietary LLM"],
        "widths": [12, 13, 13, 17],
        "fmt": "0.0000",
        "title": "Silicon Data LLM Token Expenditure Indices",
        "y_axis": "USD / M tokens",
    },
    "gpu": {
        "sheet": "GPU Rental",
        "csv": BASE / "data" / "gpu_rental_index.csv",
        "columns": ["date"] + [c for c, _, _, _ in GPU_SERIES],
        "labels": ["Date", "H100 Neo", "A100 Neo", "H200 Neo", "MI300X Neo",
                   "B200 Neo", "H100 Hyper", "A100 Hyper"],
        "widths": [12, 11, 11, 11, 12, 11, 12, 12],
        "fmt": "0.00",
        "title": "Silicon Data GPU Rental Price Indices",
        "y_axis": "USD / hr",
    },
}

# The RSC payload is escaped an unpredictable number of times; \\* absorbs any depth.
Q = r'\\*"'
RE_PAIR = re.compile(Q + r"(\d{4}-\d{2}-\d{2})" + Q + r"\s*:\s*" + Q + r"([0-9]+\.[0-9]+)" + Q)
# Anchor on the JSON key itself. A bare substring search for "indexes" matches the
# script filename "token-indexes-chart/page-*.js" in <head> first. The indexes object
# is a flat date->string map, so a non-greedy match to the first "}" is exact.
RE_INDEXES = re.compile(Q + "indexes" + Q + r"\s*:\s*\{(.*?)\}", re.S)


# Cache-busting. Observed 2026-08-31: the token endpoint served GitHub's datacenter IPs a
# response frozen at 2026-08-28 for four consecutive runs while the same code on a home
# connection got 2026-08-30. The public page served to those runners was stale in the same
# way, so portal and page agreed and the cross-check passed - two stale surfaces agreeing
# defeats a consistency check. The GPU endpoint was unaffected, and it is the one carrying
# extra query params, which is what points at an edge cache keyed on the URL.
NO_CACHE = {"Cache-Control": "no-cache", "Pragma": "no-cache"}


def bust(params=None):
    """Add a per-request token so an edge cache cannot serve a stale copy."""
    p = dict(params or {})
    p["_"] = str(int(time.time() * 1000))
    return p


# How many days behind today each feed may legitimately be before it is called stale.
# Silicon Data publishes with roughly a one-day lag, so 3 days is generous but quiet.
# Ramp is the awkward one: rows are dated to the FIRST of the month and month M is
# published around the end of M+1, so a perfectly healthy Ramp feed is routinely
# 60-90 days "old" by this measure. A first pass at 50 fired a false alert on a normal
# feed - and an alarm that cries wolf is worse than no alarm, because it trains you to
# ignore the real one. 95 still catches Ramp stopping publication outright, which is
# the only Ramp failure worth waking anyone for (it serves full history, so it
# self-heals and a missed day costs nothing).
MAX_LAG_DAYS = {"llm": 3, "gpu": 3, "fc": 3, "ramp": 95}


def check_freshness(label, latest_iso):
    """Currency check, separate from the cross-checks' consistency check."""
    try:
        lag = (date.today() - date.fromisoformat(latest_iso[:10])).days
    except ValueError:
        return None
    limit = MAX_LAG_DAYS.get(label, 3)
    if lag > limit:
        msg = f"{label}: latest {latest_iso} is {lag} days old (expected <= {limit})"
        log(f"  STALE {msg}")
        return msg
    return None


def log(msg):
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    line = f"{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%SZ')}  {msg}"
    try:
        with open(LOG_PATH, "a", encoding="utf-8") as fh:
            fh.write(line + "\n")
    except OSError:
        pass
    try:
        print(line)
    except (UnicodeEncodeError, OSError):
        pass


def telegram(text):
    """Send via the Bot API using env vars (CI), else the local sender script.

    Plain text, no parse_mode - tracker output is full of . $ [ _ and MarkdownV2
    rejects them unescaped, which fails silently from the reader's side.
    """
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat = os.environ.get("TELEGRAM_CHAT_ID")
    if token and chat:
        import urllib.parse
        import urllib.request
        try:
            body = urllib.parse.urlencode({
                "chat_id": chat, "text": text[:4000],
                "disable_web_page_preview": "true"}).encode()
            req = urllib.request.Request(
                f"https://api.telegram.org/bot{token}/sendMessage", data=body)
            with urllib.request.urlopen(req, timeout=30) as r:
                return r.status == 200
        except Exception as exc:
            log(f"  WARN telegram send failed: {type(exc).__name__}")
            return False
    if not TELEGRAM.exists():
        return False
    try:
        with tempfile.NamedTemporaryFile("w", suffix=".txt", delete=False, encoding="utf-8") as fh:
            fh.write(text)
            tmp = fh.name
        subprocess.run([sys.executable, str(TELEGRAM), "--file", tmp], timeout=60,
                       capture_output=True)
        os.unlink(tmp)
        return True
    except Exception as exc:
        log(f"  WARN telegram notify failed: {exc}")
        return False


def notify_failure(text):
    telegram(text)


def echoed(html, key):
    m = re.search(Q + re.escape(key) + Q + r"\s*:\s*" + Q + r"([A-Za-z0-9_-]+)" + Q, html)
    return m.group(1) if m else None


def parse_indexes(html, label):
    block = RE_INDEXES.search(html)
    if not block:
        raise ValueError(f"{label}: 'indexes' object not found in response")
    pairs = RE_PAIR.findall(block.group(1))
    if not pairs:
        raise ValueError(f"{label}: no date/value pairs parsed")
    out = {d: v for d, v in pairs}
    # Every entry must be a date/value pair; a leftover means the shape changed.
    if len(out) != block.group(1).count(":"):
        raise ValueError(f"{label}: indexes block has unparsed entries - format changed")
    return out


def fetch_llm(session, token):
    col, expect_name, _ = LLM_SERIES[token]
    r = session.get(TOKEN_PORTAL, params=bust({"token": token}), timeout=45, headers=NO_CACHE)
    r.raise_for_status()
    html = r.text
    if echoed(html, "token") != token:
        raise ValueError(f"{token}: server echoed token={echoed(html, 'token')!r}, not the one requested")
    if expect_name.replace(" ", "") not in html.replace("\\", "").replace(" ", ""):
        raise ValueError(f"{token}: expected index_name {expect_name!r} not found")
    out = parse_indexes(html, token)
    log(f"  {col:<16} {len(out)} days  {min(out)} -> {max(out)} (latest {out[max(out)]})")
    return col, out


def fetch_gpu(session, col, gpu, tab):
    r = session.get(GPU_PORTAL, params=bust({"standalone": "true", "gpu": gpu, "mainTab": tab}),
                    timeout=45, headers=NO_CACHE)
    r.raise_for_status()
    html = r.text
    got_gpu, got_tab = echoed(html, "gpu"), echoed(html, "initialMainTab")
    if got_gpu != gpu:
        raise ValueError(f"{col}: server echoed gpu={got_gpu!r}, not {gpu!r}")
    if got_tab != tab:
        # The server falls back to neo-cloud for tabs a GPU does not have, so an
        # unchecked mismatch would file neo-cloud prices under a hyperscaler column.
        raise ValueError(
            f"{col}: requested mainTab={tab!r} but server served {got_tab!r} - "
            f"that GPU/tab combination no longer exists; refusing to store mislabelled data")
    out = parse_indexes(html, col)
    log(f"  {col:<16} {len(out)} days  {min(out)} -> {max(out)} (latest {out[max(out)]})")
    return col, out


def fetch_marketing_page(session):
    r = session.get(MARKETING, params=bust(), timeout=45, headers=NO_CACHE)
    r.raise_for_status()
    return r.text


def parse_marketing_readings(html):
    """The rounded 2dp LLM values shown on the public page - used as a cross-check."""
    out = {}
    for key in ("llm-token", "open-llm", "proprietary-llm"):
        m = re.search(
            Q + re.escape(key) + Q + r"\s*:\s*\{" + Q + "value" + Q + r"\s*:\s*" + Q + r"([0-9.]+)" + Q
            + r".*?" + Q + "asOf" + Q + r"\s*:\s*" + Q + r"As of ([^\\\"]+)" + Q,
            html, re.S)
        if m:
            out[key] = (m.group(1), m.group(2).strip())
    return out


def parse_gpu_cards(html):
    """The 'Other Silicon Indices' cards carry each GPU's latest NEO-CLOUD price."""
    out = {}
    for gpu in ("H100", "A100", "B200", "MI300X"):
        for m in re.finditer(">" + gpu + "<", html):
            v = re.search(r"\$([0-9]+\.[0-9]{2})", html[m.end():m.end() + 1500])
            if v:
                out[gpu.lower() + "_neo"] = v.group(1)
                break
    return out


def cross_check_gpu(series, cards):
    """Card prices must match our stored value for the latest or previous day.

    Two dates are allowed because the cards publish on a slight lag behind the
    chart endpoint - observed drifting by one day. This still catches the failure
    that matters: a whole series mislabelled, swapped, or shifted by a decimal.
    """
    checked = 0
    for col, shown in cards.items():
        ours = series.get(col)
        if not ours:
            continue
        recent = sorted(ours)[-2:]
        vals = {f"{float(ours[d]):.2f}" for d in recent}
        if f"{float(shown):.2f}" not in vals:
            raise ValueError(
                f"{col}: public card shows {shown} but our last two days are "
                f"{[ours[d] for d in recent]} ({recent}) - refusing to store")
        checked += 1
    if checked < 4:
        raise ValueError(
            f"GPU cross-check covered only {checked}/4 cards - public page parse degraded, "
            f"refusing to store unverified values")
    log(f"  cross-check OK - {checked}/4 GPU neo-cloud series verified against public cards")


def cross_check(series, readings):
    """Portal 4dp values must round to the 2dp values the public page displays.

    Every series must be verified. A partial check is treated as failure: the
    marketing-page regex is the most fragile part of this script, and a degraded
    parse would otherwise let an unverified number enter a history that cannot be
    re-derived. Losing one day's capture is cheap - the 7-day window heals it.
    """
    checked, as_of = 0, None
    for col, _, mkey in LLM_SERIES.values():
        if mkey not in readings:
            log(f"  WARN no public-page reading parsed for {mkey}")
            continue
        shown, as_of = readings[mkey]
        try:
            iso = datetime.strptime(as_of, "%b %d, %Y").strftime("%Y-%m-%d")
        except ValueError:
            log(f"  WARN could not parse as-of date {as_of!r} for {mkey}")
            continue
        portal = series[col].get(iso)
        if portal is None:
            raise ValueError(f"{col}: page reports as-of {iso} but portal has no value for that date")
        if f"{float(portal):.2f}" != f"{float(shown):.2f}":
            raise ValueError(
                f"{col}: portal {portal} rounds to {float(portal):.2f} but page shows {shown} (as of {iso})")
        checked += 1
    if checked < len(LLM_SERIES):
        raise ValueError(
            f"cross-check covered only {checked}/{len(LLM_SERIES)} series - public page parse degraded, "
            f"refusing to store unverified values")
    log(f"  cross-check OK - {checked}/{len(LLM_SERIES)} LLM series verified against public page (as of {as_of})")


# -------------------------------------------------------------- forward curve

def fetch_forward_curve(session):
    """Return (as_of_date, payload) for the full 3-GPU x 145-tenor curve set."""
    r = session.get(FC_PORTAL, params=bust(), timeout=45, headers=NO_CACHE)
    r.raise_for_status()
    html = r.text
    m = re.search(Q + "data" + Q + r"\s*:\s*\{" + Q + "date" + Q, html)
    if not m:
        raise ValueError("forward curve: data payload not found")
    start = html.index("{", m.start() + 4)
    depth, i = 0, start
    while i < len(html):
        if html[i] == "{":
            depth += 1
        elif html[i] == "}":
            depth -= 1
            if depth == 0:
                i += 1
                break
        i += 1
    raw = html[start:i]
    payload = None
    for _ in range(4):
        try:
            payload = json.loads(raw)
            break
        except json.JSONDecodeError:
            raw = raw.replace('\\"', '"').replace("\\\\", "\\")
    if payload is None:
        raise ValueError("forward curve: payload would not decode")

    as_of = payload.get("date")
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", as_of or ""):
        raise ValueError(f"forward curve: bad as-of date {as_of!r}")
    for gpu in FC_GPUS:
        if gpu not in payload:
            raise ValueError(f"forward curve: {gpu} missing from payload")
        missing = [t for t in FC_TENORS if t not in payload[gpu]]
        if missing:
            raise ValueError(f"forward curve: {gpu} missing tenors {missing[:5]}")
    log(f"  forward curve     as of {as_of}  {len(FC_GPUS)} GPUs x "
        f"{len(payload[FC_GPUS[0]])} tenors x {len(FC_RATES)} rates")
    return as_of, payload


def archive_raw_curve(as_of, payload):
    """Keep the untouched payload - all 145 tenors - so 'in totality' is literal."""
    FC_RAW_DIR.mkdir(parents=True, exist_ok=True)
    dest = FC_RAW_DIR / f"{as_of}.json"
    if dest.exists():
        return False
    fd, tmp = tempfile.mkstemp(dir=FC_RAW_DIR, suffix=".tmp")
    with os.fdopen(fd, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, separators=(",", ":"), sort_keys=True)
    os.replace(tmp, dest)
    return True


def curve_rows(as_of, payload):
    """Flatten to one row per (date, gpu, rate_type); tenors become columns."""
    rows = {}
    for gpu in FC_GPUS:
        for rate in FC_RATES:
            row = {"date": as_of, "gpu": gpu, "rate_type": rate}
            for t in FC_TENORS:
                v = payload[gpu][t].get(rate)
                if v is None:
                    raise ValueError(f"forward curve: {gpu} tenor {t} missing {rate}")
                row[f"t{t}"] = f"{float(v):.4f}"
            rows[(as_of, gpu, rate)] = row
    return rows


def cross_check_curve(payload, gpu_series, as_of):
    """The curve's spot point must agree with the separately-scraped spot index.

    Two independent endpoints, one number - the strongest check available here.
    Also asserts term == forward at tenor 0, which is true by construction.
    """
    checked = 0
    for gpu in FC_GPUS:
        node = payload[gpu]["0"]
        term, fwd = float(node["term_rate"]), float(node["forward_rate"])
        if f"{term:.4f}" != f"{fwd:.4f}":
            raise ValueError(f"forward curve {gpu}: tenor 0 term {term} != forward {fwd}")
        spot = (gpu_series or {}).get(f"{gpu.lower()}_neo", {}).get(as_of)
        if spot is None:
            continue
        if f"{term:.2f}" != f"{float(spot):.2f}":
            raise ValueError(
                f"forward curve {gpu}: spot node {term} disagrees with scraped "
                f"neo-cloud index {spot} on {as_of}")
        checked += 1
    if gpu_series and checked < len(FC_GPUS):
        log(f"  WARN curve spot cross-check covered only {checked}/{len(FC_GPUS)} GPUs "
            f"(spot index has no {as_of} row yet)")
    else:
        log(f"  cross-check OK - {checked}/{len(FC_GPUS)} curve spot nodes match the spot index")


def load_fc_csv():
    if not FC_CSV.exists():
        return {}
    with open(FC_CSV, newline="", encoding="utf-8") as fh:
        return {(r["date"], r["gpu"], r["rate_type"]): r for r in csv.DictReader(fh)}


def write_fc_csv(rows):
    FC_CSV.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=FC_CSV.parent, suffix=".tmp")
    with os.fdopen(fd, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=FC_COLUMNS)
        w.writeheader()
        for k in sorted(rows):
            w.writerow({c: rows[k].get(c, "") for c in FC_COLUMNS})
    os.replace(tmp, FC_CSV)


def _sync_fc_sheet(wb, rows, rebuild):
    """Append curve snapshots not already present. Never rewrites a stored curve."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if rebuild and FC_SHEET in wb.sheetnames:
        del wb[FC_SHEET]
    created = FC_SHEET not in wb.sheetnames
    if created:
        ws = wb.create_sheet(FC_SHEET)
        headers = ["Date", "GPU", "Rate"] + [f"{t}m" for t in FC_TENORS]
        ws.append(headers)
        fill = PatternFill("solid", fgColor="1F2933")
        thin = Side(style="thin", color="D0D5DA")
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        ws.freeze_panes = "D2"
        ws.row_dimensions[1].height = 22
        for c, w in ((1, 12), (2, 9), (3, 14)):
            ws.column_dimensions[get_column_letter(c)].width = w
        for i in range(len(FC_TENORS)):
            ws.column_dimensions[get_column_letter(4 + i)].width = 8
    else:
        ws = wb[FC_SHEET]

    present = set()
    for r in range(2, ws.max_row + 1):
        d, g, rt = (ws.cell(row=r, column=c).value for c in (1, 2, 3))
        if d is None:
            continue
        d = d.strftime("%Y-%m-%d") if isinstance(d, datetime) else str(d)[:10]
        present.add((d, str(g), str(rt)))

    row_at = ws.max_row
    added = 0
    for key in sorted(rows):
        if key in present:
            continue
        row_at += 1
        rec = rows[key]
        ws.cell(row=row_at, column=1,
                value=datetime.strptime(rec["date"], "%Y-%m-%d")).number_format = "yyyy-mm-dd"
        ws.cell(row=row_at, column=2, value=rec["gpu"])
        ws.cell(row=row_at, column=3, value=rec["rate_type"])
        for i, t in enumerate(FC_TENORS):
            v = rec.get(f"t{t}")
            cell = ws.cell(row=row_at, column=4 + i, value=float(v) if v else None)
            cell.number_format = "0.0000"
        added += 1
    ws.auto_filter.ref = f"A1:{get_column_letter(3 + len(FC_TENORS))}{ws.max_row}"
    return added, ("rebuilt" if rebuild else "created" if created else "appended")


# ------------------------------------------------------------- Ramp AI Index

def _rsc_stream(html):
    """Concatenate the payload strings from every self.__next_f.push([1,"..."]).

    The series are JSON inside a JS string literal, so they must be unescaped by
    decoding that literal - not by chewing backslashes, which silently fails.
    """
    out = []
    for m in re.finditer(r"self\.__next_f\.push\(", html):
        i = html.index("(", m.start()) + 1
        j = _match_bracket(html, i)
        if j is None:
            continue
        try:
            arr = json.loads(html[i:j])
        except json.JSONDecodeError:
            continue
        if isinstance(arr, list) and len(arr) > 1 and isinstance(arr[1], str):
            out.append(arr[1])
    return "".join(out)


def _match_bracket(text, start):
    """Return the index just past the bracket group opening at `start`."""
    depth, i, instr, esc = 0, start, False, False
    while i < len(text):
        c = text[i]
        if instr:
            if esc:
                esc = False
            elif c == "\\":
                esc = True
            elif c == '"':
                instr = False
        else:
            if c == '"':
                instr = True
            elif c in "[({":
                depth += 1
            elif c in "])}":
                depth -= 1
                if depth == 0:
                    return i + 1
        i += 1
    return None


def _extract_array(stream, key):
    m = re.search(r'"' + key + r'"\s*:\s*\[', stream)
    if not m:
        raise ValueError(f"ramp: array {key!r} not found")
    start = stream.index("[", m.end() - 1)
    end = _match_bracket(stream, start)
    if end is None:
        raise ValueError(f"ramp: array {key!r} not terminated")
    return json.loads(stream[start:end])


def _num(v):
    """Ramp emits '$undefined' for suppressed cells."""
    return None if v is None or isinstance(v, str) else v


def fetch_ramp(session):
    r = session.get(RAMP_URL, params=bust(), timeout=90, headers=NO_CACHE)
    r.raise_for_status()
    stream = _rsc_stream(r.text)
    if len(stream) < 100_000:
        raise ValueError(f"ramp: RSC stream only {len(stream)} chars - page structure changed")
    needed = {src for _, src, _, _ in RAMP_ADOPTION + RAMP_SPEND}
    arrays = {k: _extract_array(stream, k) for k in needed}
    for k, arr in arrays.items():
        if not arr:
            raise ValueError(f"ramp: array {k!r} is empty")
    return arrays


def build_ramp_rows(arrays, which):
    """Flatten to one row per (month, breakdown, dimension). Long format so the
    ragged breakdowns share a sheet and pivot cleanly."""
    spec = RAMP_SHEETS[which]
    rows = {}
    for label, src, dim_field, dim_type in spec["specs"]:
        for rec in arrays[src]:
            if dim_type and rec.get("dimension_type") != dim_type:
                continue
            dim = str(rec.get(dim_field, "")) if dim_field else ""
            key = (rec["date_month"][:10], label, dim)
            row = {"date_month": key[0], "breakdown": label, "dimension": dim}
            for metric in spec["metrics"]:
                v = _num(rec.get(metric))
                row[metric] = "" if v is None else f"{float(v):.4f}"
            rows[key] = row
    return rows


def cross_check_ramp(arrays):
    """Structural invariants. Cheap, and they catch a silently reshaped feed."""
    months = {}
    for label, src, _, _ in RAMP_ADOPTION + RAMP_SPEND:
        months[src] = max(r["date_month"] for r in arrays[src])
    head = arrays["adoptionOverall"]
    rates = [r["adoption_rate_pct"] for r in head]
    if not all(0 <= x <= 100 for x in rates):
        raise ValueError(f"ramp: adoption rate outside 0-100 ({min(rates)}..{max(rates)})")
    # mom_change_pp must equal the actual month-on-month level difference
    ordered = sorted(head, key=lambda r: r["date_month"])
    for prev, cur in zip(ordered, ordered[1:]):
        stated, actual = cur.get("mom_change_pp"), cur["adoption_rate_pct"] - prev["adoption_rate_pct"]
        if stated is not None and abs(stated - actual) > 0.05:
            raise ValueError(
                f"ramp: {cur['date_month']} mom_change_pp {stated} != level diff {actual:.2f}")
    for which in RAMP_SHEETS:
        for label, src, _, dim_type in RAMP_SHEETS[which]["specs"]:
            n = sum(1 for r in arrays[src] if not dim_type or r.get("dimension_type") == dim_type)
            if n == 0:
                raise ValueError(f"ramp: breakdown {label!r} ({src}/{dim_type}) has no records")
    log(f"  cross-check OK - ramp structure valid, latest month "
        f"{max(months.values())} across {len(months)} arrays")
    return max(months.values())


def archive_raw_ramp(arrays, latest_month):
    """One vintage per published month. Ramp restates history, and Citadel's
    figures already differ slightly from ours, so keeping vintages matters."""
    RAMP_RAW_DIR.mkdir(parents=True, exist_ok=True)
    dest = RAMP_RAW_DIR / f"{latest_month}.json"
    if dest.exists():
        return False
    fd, tmp = tempfile.mkstemp(dir=RAMP_RAW_DIR, suffix=".tmp")
    with os.fdopen(fd, "w", encoding="utf-8") as fh:
        json.dump(arrays, fh, separators=(",", ":"), sort_keys=True)
    os.replace(tmp, dest)
    return True


def load_keyed_csv(path, key_fields):
    if not path.exists():
        return {}
    with open(path, newline="", encoding="utf-8") as fh:
        return {tuple(r[k] for k in key_fields): r for r in csv.DictReader(fh)}


def write_keyed_csv(path, columns, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=path.parent, suffix=".tmp")
    with os.fdopen(fd, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=columns)
        w.writeheader()
        for k in sorted(rows):
            w.writerow({c: rows[k].get(c, "") for c in columns})
    os.replace(tmp, path)


def _sync_ramp_sheet(wb, which, rows, rebuild):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    spec = RAMP_SHEETS[which]
    name = spec["sheet"]
    if rebuild and name in wb.sheetnames:
        del wb[name]
    created = name not in wb.sheetnames
    if created:
        ws = wb.create_sheet(name)
        ws.append(spec["labels"])
        fill = PatternFill("solid", fgColor="1F2933")
        thin = Side(style="thin", color="D0D5DA")
        for c in range(1, len(spec["labels"]) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
        ws.freeze_panes = "D2"
        ws.row_dimensions[1].height = 30
        for i, w in enumerate(spec["widths"]):
            ws.column_dimensions[get_column_letter(1 + i)].width = w
    else:
        ws = wb[name]

    present = set()
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        if d is None:
            continue
        d = d.strftime("%Y-%m-%d") if isinstance(d, datetime) else str(d)[:10]
        present.add((d, str(ws.cell(row=r, column=2).value or ""),
                     str(ws.cell(row=r, column=3).value or "")))

    row_at, added = ws.max_row, 0
    for key in sorted(rows):
        if key in present:
            continue
        row_at += 1
        rec = rows[key]
        ws.cell(row=row_at, column=1,
                value=datetime.strptime(rec["date_month"], "%Y-%m-%d")).number_format = "yyyy-mm"
        ws.cell(row=row_at, column=2, value=rec["breakdown"])
        ws.cell(row=row_at, column=3, value=rec["dimension"])
        for i, metric in enumerate(spec["metrics"]):
            v = rec.get(metric)
            cell = ws.cell(row=row_at, column=4 + i, value=float(v) if v not in (None, "") else None)
            cell.number_format = spec["fmt"]
        added += 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(spec['labels']))}{ws.max_row}"
    return added, ("rebuilt" if rebuild else "created" if created else "appended")


# ------------------------------------------------------------------ archive io

def backup(path):
    """Snapshot a file before modifying it. History here is irreplaceable."""
    if not path.exists():
        return
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = BACKUP_DIR / f"{path.stem}_{stamp}{path.suffix}"
    try:
        shutil.copy2(path, dest)
    except OSError as exc:
        log(f"  WARN backup of {path.name} failed: {exc}")
        return
    for f in sorted(BACKUP_DIR.glob(f"{path.stem}_*{path.suffix}"))[:-KEEP_BACKUPS]:
        try:
            f.unlink()
        except OSError:
            pass


def load_csv(path):
    if not path.exists():
        return {}
    with open(path, newline="", encoding="utf-8") as fh:
        return {row["date"]: row for row in csv.DictReader(fh)}


def write_csv(path, columns, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=path.parent, suffix=".tmp")
    with os.fdopen(fd, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=columns)
        w.writeheader()
        for d in sorted(rows):
            w.writerow({c: rows[d].get(c, "") for c in columns})
    os.replace(tmp, path)


def upsert(existing, series):
    added, revised = [], []
    dates = set()
    for col_map in series.values():
        dates |= set(col_map)
    for d in sorted(dates):
        row = existing.get(d)
        if row is None:
            row = {"date": d}
            existing[d] = row
            added.append(d)
        for col, values in series.items():
            new = values.get(d)
            if new is None:
                continue
            old = (row.get(col) or "").strip()
            if old and old != new:
                # Provider restated a published number - never swap it silently.
                revised.append(f"{d} {col}: {old} -> {new}")
            row[col] = new
    return added, revised


# ------------------------------------------------------------------- workbook

def _write_row(ws, r, date_str, row, spec):
    """Write one data row. Only ever touches the machine-owned columns."""
    from openpyxl.utils import get_column_letter

    cols = spec["columns"][1:]
    n = len(cols)
    ws.cell(row=r, column=1, value=datetime.strptime(date_str, "%Y-%m-%d")).number_format = "yyyy-mm-dd"
    for i, col in enumerate(cols):
        v = row.get(col)
        cell = ws.cell(row=r, column=2 + i, value=float(v) if v else None)
        cell.number_format = spec["fmt"]
    if r > 2:  # DoD % - blank on the very first data row, nothing to compare against
        for i in range(n):
            src = get_column_letter(2 + i)
            cell = ws.cell(row=r, column=2 + n + i,
                           value=f"=IF(OR({src}{r}=\"\",{src}{r-1}=\"\"),\"\",{src}{r}/{src}{r-1}-1)")
            cell.number_format = "0.00%"


def _create_sheet(wb, spec):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(spec["sheet"])
    headers = spec["labels"] + [f"{l} DoD %" for l in spec["labels"][1:]]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1F2933")
    thin = Side(style="thin", color="D0D5DA")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 30
    for i, w in enumerate(spec["widths"]):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    for i in range(len(spec["columns"]) - 1):
        ws.column_dimensions[get_column_letter(1 + len(spec["columns"]) + i)].width = 14
    return ws


def _sync_sheet(wb, spec, rows, rebuild):
    """Append missing dates to one dataset's sheet. Returns (n_appended, mode)."""
    ordered = sorted(rows)
    name = spec["sheet"]

    if rebuild and name in wb.sheetnames:
        del wb[name]
    if name not in wb.sheetnames:
        ws = _create_sheet(wb, spec)
        for i, d in enumerate(ordered):
            _write_row(ws, i + 2, d, rows[d], spec)
        return len(ordered), "rebuilt" if rebuild else "created"

    ws = wb[name]
    present = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, datetime):
            present.add(v.strftime("%Y-%m-%d"))
        elif isinstance(v, str) and v.strip():
            present.add(v.strip()[:10])
    missing = [d for d in ordered if d not in present]
    # Appending assumes dates arrive in order. An older date would misorder the
    # DoD chain, so regenerate that sheet from the archive instead.
    if missing and present and min(missing) < max(present):
        log(f"  note: out-of-order date on '{name}', regenerating from archive")
        return _sync_sheet(wb, spec, rows, rebuild=True)
    row_at = ws.max_row
    for d in missing:
        row_at += 1
        _write_row(ws, row_at, d, rows[d], spec)
    ws.auto_filter.ref = (f"A1:{ws.cell(row=1, column=2 * len(spec['columns']) - 1).coordinate[:-1]}"
                          f"{len(ordered) + 1}")
    return len(missing), "appended"


def _rebuild_chart_sheet(wb, store):
    """Machine-owned and regenerated, so refs always span the full history.
    Charts you build yourself belong on your own sheet."""
    from openpyxl.chart import LineChart, Reference

    if CHART_SHEET in wb.sheetnames:
        del wb[CHART_SHEET]
    cs = wb.create_sheet(CHART_SHEET)
    anchor_row = 2
    for key, spec in DATASETS.items():
        rows = store.get(key)
        if not rows or spec["sheet"] not in wb.sheetnames:
            continue
        last = len(rows) + 1
        if last < 3:
            continue
        data = wb[spec["sheet"]]
        chart = LineChart()
        chart.title = spec["title"]
        chart.y_axis.title = spec["y_axis"]
        chart.x_axis.title = "Date"
        chart.height, chart.width = 12, 32
        chart.add_data(Reference(data, min_col=2, max_col=len(spec["columns"]),
                                 min_row=1, max_row=last), titles_from_data=True)
        chart.set_categories(Reference(data, min_col=1, min_row=2, max_row=last))
        cs.add_chart(chart, f"B{anchor_row}")
        anchor_row += 25


def _rebuild_source_sheet(wb, store, curves=None, ramp=None):
    from openpyxl.styles import Font

    if SOURCE_SHEET in wb.sheetnames:
        del wb[SOURCE_SHEET]
    meta = wb.create_sheet(SOURCE_SHEET)
    lines = [
        ["Source", "Silicon Data - silicondata.com"],
        ["Update mode", "Append-only. New dates are added as new rows; existing rows are never rewritten."],
        ["Backups", f"{BACKUP_DIR} (last {KEEP_BACKUPS} of each file)"],
        ["", ""],
        ["SHEET", "'Daily Index' - LLM token expenditure indices"],
        ["Endpoint", f"{TOKEN_PORTAL}?token={{expenditure|open_expenditure|closed_expenditure}}"],
        ["Units", "USD per million tokens, usage-weighted blended rate"],
        ["Precision", "4dp from the portal endpoint (the public page rounds to 2dp)"],
        ["Machine cols", "A:G - do not hand-edit.  YOURS: H onward"],
        ["", ""],
        ["SHEET", "'GPU Rental' - GPU rental price indices"],
        ["Endpoint", f"{GPU_PORTAL}?standalone=true&gpu={{h100|a100|h200|mi300x|b200}}&mainTab={{neo-cloud|hyperscaler}}"],
        ["Units", "USD per GPU-hour, spot rental"],
        ["Precision", "2dp as published"],
        ["Coverage", "Hyperscaler exists for H100 and A100 ONLY. B200/MI300X/H200 are neo-cloud only - "
                     "the server silently returns neo-cloud data if you ask for their hyperscaler tab."],
        ["H200 note", "Hidden on the public site (visible only in standalone mode) but the series is real."],
        ["Tickers", "SDH100RT, SDA100RT, SDMI300XRT, SDB200RT (Bloomberg, neo-cloud)"],
        ["Machine cols", "A:O - do not hand-edit.  YOURS: P onward"],
        ["", ""],
        ["SHEET", "'Forward Curve' - GPU rental term structure and forward prices"],
        ["Endpoint", FC_PORTAL],
        ["Shape", "ONE ROW = ONE COMPLETE CURVE. Filter Date/GPU/Rate, then plot across the tenor "
                  "columns to overlay today vs a week ago vs a month ago."],
        ["Units", "USD per GPU-hour. Neo-cloud only. Tenor 0m = spot."],
        ["Tenors", "0 to 36 months. The source also emits 108 interpolated quarter-month points; "
                   "those are kept verbatim in the raw JSON archive, not here."],
        ["Raw archive", f"{FC_RAW_DIR} - full 145-tenor payload, one file per day, never deleted"],
        ["Snapshots", "One curve per day, no history and no backfill at source. A day not "
                      "captured is lost permanently - unlike the 7-day series, this has NO self-heal."],
        ["Machine cols", "A:AN - do not hand-edit.  YOURS: AO onward"],
        ["", ""],
        ["Coverage limit", "The token and GPU endpoints serve a rolling 7-day window only - history "
                           "exists ONLY in this file and the CSV archives. Gaps up to 6 days self-heal; "
                           "longer is unrecoverable. The forward curve has no window at all."],
    ]
    lines += [
        ["", ""],
        ["SHEETS", "'Ramp Adoption' and 'Ramp Spend per Employee' - Ramp AI Index"],
        ["Source", RAMP_URL],
        ["What", "Monthly AI adoption and spend from card/bill-pay transactions across 70,000+ US firms"],
        ["Shape", "Long format: Month | Breakdown | Dimension | metrics. Filter Breakdown to get "
                  "Headline / Sector / Business size / Models (adoption) or Overall / Business size / "
                  "Sector / Financing status (spend)."],
        ["Cadence", "MONTHLY, published with a lag. Checked daily; a new month is appended when it "
                    "appears, otherwise nothing changes."],
        ["Self-healing", "Full history is served on every request, so unlike the Silicon Data feeds "
                         "a missed day costs nothing here."],
        ["Vintages", f"{RAMP_RAW_DIR} - one raw snapshot per published month. Ramp restates history."],
        ["Caution", "Citadel's 'ratio of 90th-percentile to median' uses top_10_percent_median_pepm "
                    "(Top 10% median $ here), NOT the P90 $ column. 650/11.95 = 54x."],
        ["Verified", "Field mapping reproduces Citadel's published July-2026 moves: top 1% +48.8% "
                     "(printed 49%), top 10% +24.9% (25%), median +9.3% (9%)."],
        ["Machine cols", "'Ramp Adoption' A:F, 'Ramp Spend per Employee' A:I - do not hand-edit"],
    ]
    if curves:
        dates = sorted({k[0] for k in curves})
        lines.append(["Archive (fc)", str(FC_CSV)])
        lines.append(["Span (fc)", f"{dates[0]} to {dates[-1]}  ({len(dates)} daily snapshots, "
                                   f"{len(curves)} curve rows)"])
    for which, rows in (ramp or {}).items():
        months = sorted({k[0] for k in rows})
        lines.append([f"Archive ({which})", str(RAMP_SHEETS[which]["csv"])])
        lines.append([f"Span ({which})", f"{months[0]} to {months[-1]}  ({len(months)} months, "
                                         f"{len(rows)} rows)" if months else "empty"])
    for key, spec in DATASETS.items():
        rows = store.get(key) or {}
        o = sorted(rows)
        lines.append([f"Archive ({key})", str(spec["csv"])])
        lines.append([f"Span ({key})", f"{o[0]} to {o[-1]}  ({len(o)} rows)" if o else "empty"])
    lines.append(["Last updated", datetime.now().strftime("%Y-%m-%d %H:%M local")])
    for row in lines:
        meta.append(row)
    meta.column_dimensions["A"].width = 16
    meta.column_dimensions["B"].width = 120
    for r in range(1, meta.max_row + 1):
        meta.cell(row=r, column=1).font = Font(bold=True)


def update_xlsx(store, curves=None, ramp=None, rebuild=False):
    """Sync every dataset's sheet in a single load/save of the workbook."""
    from openpyxl import Workbook, load_workbook

    if rebuild or not XLSX_PATH.exists():
        wb = Workbook()
        wb.remove(wb.active)
    else:
        wb = load_workbook(XLSX_PATH)

    results = {}
    for key, spec in DATASETS.items():
        if key in store:
            results[key] = _sync_sheet(wb, spec, store[key], rebuild)
    if curves:
        results["fc"] = _sync_fc_sheet(wb, curves, rebuild)
    for which, rows in (ramp or {}).items():
        results[which] = _sync_ramp_sheet(wb, which, rows, rebuild)

    _rebuild_chart_sheet(wb, store)
    _rebuild_source_sheet(wb, store, curves, ramp)
    # Keep machine sheets first, user sheets after, without reordering user work.
    order = [s for s in (list(d["sheet"] for d in DATASETS.values()) + [FC_SHEET]
                         + [s["sheet"] for s in RAMP_SHEETS.values()]
                         + [CHART_SHEET, SOURCE_SHEET])
             if s in wb.sheetnames]
    wb._sheets = ([wb[s] for s in order] + [s for s in wb._sheets if s.title not in order])

    fd, tmp = tempfile.mkstemp(dir=XLSX_PATH.parent, suffix=".xlsx")
    os.close(fd)
    wb.save(tmp)
    os.replace(tmp, XLSX_PATH)
    return results


# ------------------------------------------------------------------------ main

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-xlsx", action="store_true", help="update the CSV archives only")
    ap.add_argument("--quiet", action="store_true", help="suppress the failure Telegram")
    ap.add_argument("--rebuild", action="store_true",
                    help="regenerate the whole workbook from the CSV archives "
                         "(DISCARDS your own sheets and columns - recovery use only)")
    ap.add_argument("--only", choices=sorted(list(DATASETS) + ["fc", "ramp"]),
                    help="run a single dataset (llm, gpu, fc, or ramp)")
    args = ap.parse_args()

    wanted = [args.only] if args.only else list(DATASETS) + ["fc", "ramp"]
    log(f"=== run start ({', '.join(wanted)})")
    session = requests.Session()
    session.headers.update({"User-Agent": UA, "Accept-Language": "en-US,en;q=0.9"})

    store = {}
    page = fetch_marketing_page(session)  # one fetch, cross-checks both datasets

    if "llm" in wanted:
        series = {}
        for token in LLM_SERIES:
            col, values = fetch_llm(session, token)
            series[col] = values
        cross_check(series, parse_marketing_readings(page))
        store["llm"] = series

    if "gpu" in wanted:
        series = {}
        for col, gpu, tab, _ in GPU_SERIES:
            col, values = fetch_gpu(session, col, gpu, tab)
            series[col] = values
        cross_check_gpu(series, parse_gpu_cards(page))
        store["gpu"] = series

    curves = None
    if "fc" in wanted:
        as_of, payload = fetch_forward_curve(session)
        cross_check_curve(payload, store.get("gpu"), as_of)
        if archive_raw_curve(as_of, payload):
            log(f"  raw curve archived: {as_of}.json (all 145 tenors)")
        fresh = curve_rows(as_of, payload)
        curves = load_fc_csv()
        before = len(curves)
        new = [k for k in fresh if k not in curves]
        for k in new:
            curves[k] = fresh[k]
        if len(curves) < before:
            raise ValueError(f"fc: refusing to write, archive shrank {before} -> {len(curves)}")
        backup(FC_CSV)
        write_fc_csv(curves)
        snaps = len({k[0] for k in curves})
        log(f"  fc csv: {len(curves)} curve rows across {snaps} snapshot(s), {len(new)} new")

    ramp = None
    if "ramp" in wanted:
        arrays = fetch_ramp(session)
        latest_month = cross_check_ramp(arrays)
        if archive_raw_ramp(arrays, latest_month):
            log(f"  raw ramp vintage archived: {latest_month}.json")
        ramp = {}
        for which, spec in RAMP_SHEETS.items():
            fresh = build_ramp_rows(arrays, which)
            cols = ["date_month", "breakdown", "dimension"] + spec["metrics"]
            rows = load_keyed_csv(spec["csv"], ["date_month", "breakdown", "dimension"])
            before = len(rows)
            new, revised = [], []
            for k, rec in fresh.items():
                old = rows.get(k)
                if old is None:
                    rows[k] = rec
                    new.append(k)
                    continue
                for m in spec["metrics"]:
                    a, b = (old.get(m) or "").strip(), rec[m]
                    if a and b and a != b:
                        # Ramp restates history - surface it, never swap silently.
                        revised.append(f"{k[0]} {k[1]}/{k[2] or '-'} {m}: {a} -> {b}")
                rows[k] = rec
            if len(rows) < before:
                raise ValueError(f"{which}: refusing to write, archive shrank {before} -> {len(rows)}")
            backup(spec["csv"])
            write_keyed_csv(spec["csv"], cols, rows)
            months = sorted({k[0] for k in rows})
            log(f"  {which} csv: {len(rows)} rows, {len(months)} months "
                f"({months[0]}..{months[-1]}), {len(new)} new")
            for r in revised[:20]:
                log(f"  REVISED [{which}] {r}")
            if len(revised) > 20:
                log(f"  REVISED [{which}] ... and {len(revised) - 20} more")
            ramp[which] = rows

    archives = {}
    for key in [k for k in wanted if k in DATASETS]:
        spec = DATASETS[key]
        rows = load_csv(spec["csv"])
        before = len(rows)
        added, revised = upsert(rows, store[key])
        if len(rows) < before:
            raise ValueError(f"{key}: refusing to write, archive shrank {before} -> {len(rows)} rows")
        backup(spec["csv"])
        write_csv(spec["csv"], spec["columns"], rows)
        log(f"  {key} csv: {len(rows)} rows total, {len(added)} new {added if added else ''}")
        for r in revised:
            log(f"  REVISED [{key}] {r}")
        archives[key] = rows

    if not args.no_xlsx:
        try:
            backup(XLSX_PATH)
            for key, (n, mode) in update_xlsx(archives, curves, ramp, rebuild=args.rebuild).items():
                if key == "fc":
                    sheet, total = FC_SHEET, len(curves)
                elif key in RAMP_SHEETS:
                    sheet, total = RAMP_SHEETS[key]["sheet"], len(ramp[key])
                else:
                    sheet, total = DATASETS[key]["sheet"], len(archives[key])
                log(f"  xlsx '{sheet}' {mode}: {n} row(s), {total} total")
        except PermissionError:
            # Workbook open in Excel. CSVs already have the data; next run catches up.
            log("  WARN xlsx locked (open in Excel?) - CSVs updated, sheets sync next run")

    # Currency check, separate from the cross-checks. Those verify the portal agrees
    # with the public page; they cannot catch BOTH surfaces being stale together, which
    # is exactly what an edge cache did to the token feed on 2026-08-28..30.
    stale = []
    for key, rows in archives.items():
        if rows:
            stale.append(check_freshness(key, max(rows)))
    if curves:
        stale.append(check_freshness("fc", max(k[0] for k in curves)))
    if ramp:
        stale.append(check_freshness("ramp", max(k[0] for k in next(iter(ramp.values())))))
    stale = [s for s in stale if s]
    if stale:
        body = "\n".join("- " + s for s in stale)
        notify_failure(
            "AI Compute Tape - STALE FEED\n\n" + body
            + "\n\nThe run completed and wrote what it received, but at least one source is "
              "serving old data. Check whether that endpoint is being edge-cached.")

    tail = "  ".join(f"{k}={max(v)}" for k, v in archives.items() if v)
    if curves:
        tail += f"  fc={max(k[0] for k in curves)}"
    if ramp:
        tail += f"  ramp={max(k[0] for k in next(iter(ramp.values())))}"
    log("=== run " + ("ok" if not stale else f"ok (WITH {len(stale)} STALE FEED(S))") + "  " + tail)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:
        log(f"=== run FAILED: {type(exc).__name__}: {exc}")
        if "--quiet" not in sys.argv:
            notify_failure(f"Silicon Data scrape FAILED\n\n{type(exc).__name__}: {exc}\n\nLog: {LOG_PATH}")
        sys.exit(1)
